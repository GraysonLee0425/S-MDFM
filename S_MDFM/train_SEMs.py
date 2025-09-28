import torch
import openpyxl
from torch import nn, optim
from torch.utils.data import DataLoader
from Dataloder_chain import MyDataset
from model_chain import resnet18
import time
import os
import numpy as np

device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

def create_paths(stage):
    base_path = f"F:/wq-data/final/{stage}-20/"
    img_paths = [
        f"{base_path}rgb/",
        f"{base_path}blue/",
        f"{base_path}green/",
        f"{base_path}red/",
        f"{base_path}rededge/",
        f"{base_path}nir/",
        f"{base_path}3T/"
    ]
    label_path_train = os.path.abspath(f"Dataset/{stage}_20/{stage}_20_ndyield_train.json")
    label_path_val = os.path.abspath(f"Dataset/{stage}_20/{stage}_20_ndyield_val.json")
    excel_paths = os.path.abspath(f"{stage}_20_nd-yield.csv")

    return img_paths, label_path_train, label_path_val, excel_paths, stage


model = resnet18(num_classes=1, include_top=True, text_feature_dim=18).to(device)
criterion = nn.MSELoss()
optimizer = optim.Adam(model.parameters(), lr=0.0001)
num_params = sum(p.numel() for p in model.parameters() if p.requires_grad)

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "Metrics"
sheet1.append(["Epoch", "Train Loss", "Train R2", "Train MAE", "Train RMSE",
               "Val Loss", "Val R2", "Val MAE", "Val RMSE"])

sheet2 = workbook.create_sheet(title="Predictions")
sheet2.append(["True Values", "Predicted Values"])

workbook_para = openpyxl.Workbook()
sheet_para = workbook_para.active
sheet_para.title = "Training Parameters"
sheet_para.append(["Batch Size (Train)", "Batch Size (Val)", "Number of Parameters",
                   "Training Time (s)", "Convergence Status"])

best_val_loss = float('inf')
best_r2 = float('-inf')
num_epochs = 500
sheet_para.append([16, 8, num_params, "", ""])


def train_model():
    global best_val_loss, best_r2
    final_val_outputs = []
    final_val_labels = []

    stage_train_outputs = []
    stage_train_labels = []
    stage_sems = []  # 存放每个阶段的 SEMs

    for stage in range(5):  # 五个阶段
        img_paths, label_path_train, label_path_val, excel_paths, stage = create_paths(f"{stage+1}")

        stage = int(stage)
        TrainDataset = MyDataset(
            img_paths[0], img_paths[1], img_paths[2], img_paths[3],
            img_paths[4], img_paths[5], img_paths[6],
            label_path_train, excel_paths, flag=True, stage=stage
        )
        ValDataset = MyDataset(
            img_paths[0], img_paths[1], img_paths[2], img_paths[3],
            img_paths[4], img_paths[5], img_paths[6],
            label_path_val, excel_paths, flag=False, stage=stage
        )

        train_loader = DataLoader(TrainDataset, batch_size=16, shuffle=True,
                                  num_workers=4, pin_memory=True)
        val_loader = DataLoader(ValDataset, batch_size=8, shuffle=False,
                                num_workers=4, pin_memory=True)

        model = resnet18(num_classes=1, include_top=True, text_feature_dim=18).to(device)
        criterion = nn.MSELoss()
        optimizer = optim.Adam(model.parameters(), lr=0.0001)
        model = model.to(device)

        # 记录最后100 epoch的指标
        train_losses, val_rmses, val_maes = [], [], []

        for epoch in range(num_epochs):
            start_time = time.time()

            model.train()
            running_loss = 0.0
            all_train_outputs, all_train_labels = [], []

            for inputs, text_features, labels in train_loader:
                inputs, text_features, labels = inputs.float().to(device), text_features.float().to(device), labels.float().unsqueeze(1).to(device)

                optimizer.zero_grad()
                if stage == 0:
                    outputs = model(inputs, text_features, sems=None)
                else:
                    outputs = model(inputs, text_features, sems=stage_sems[-1].to(device))

                loss = criterion(outputs, labels)
                loss.backward()
                optimizer.step()

                running_loss += loss.item() * inputs.size(0)
                all_train_outputs.append(outputs.detach().cpu())
                all_train_labels.append(labels.detach().cpu())

            train_loss = running_loss / len(train_loader.dataset)
            all_train_outputs = torch.cat(all_train_outputs)
            all_train_labels = torch.cat(all_train_labels)
            train_r2 = 1 - ((all_train_outputs - all_train_labels).pow(2).sum() /
                            (all_train_labels - all_train_labels.mean()).pow(2).sum()).item()
            train_mae = torch.abs(all_train_outputs - all_train_labels).mean().item()
            train_rmse = torch.sqrt(torch.mean((all_train_outputs - all_train_labels) ** 2)).item()

            model.eval()
            val_loss = 0.0
            all_outputs, all_labels = [], []
            with torch.no_grad():
                for inputs, text_features, labels in val_loader:
                    inputs, text_features, labels = inputs.float().to(device), text_features.float().to(device), labels.float().unsqueeze(1).to(device)
                    if stage == 0:
                        outputs = model(inputs, text_features, sems=None)
                    else:
                        outputs = model(inputs, text_features, sems=stage_sems[-1].to(device))

                    loss = criterion(outputs, labels)
                    val_loss += loss.item() * inputs.size(0)
                    all_outputs.append(outputs.cpu())
                    all_labels.append(labels.cpu())

            val_loss = val_loss / len(val_loader.dataset)
            all_outputs = torch.cat(all_outputs)
            all_labels = torch.cat(all_labels)
            r2 = 1 - ((all_outputs - all_labels).pow(2).sum() /
                      (all_labels - all_labels.mean()).pow(2).sum()).item()
            mae = torch.abs(all_outputs - all_labels).mean().item()
            rmse = torch.sqrt(torch.mean((all_outputs - all_labels) ** 2)).item()

            # 记录最后100 epoch的指标
            if epoch >= num_epochs - 100:
                train_losses.append(train_loss)
                val_rmses.append(rmse)
                val_maes.append(mae)

            sheet1.append([epoch + 1, train_loss, train_r2, train_mae, train_rmse,
                           val_loss, r2, mae, rmse])

            if val_loss < best_val_loss:
                best_val_loss = val_loss
                torch.save(model.state_dict(), f'ChainHTF/{stage+1}_{epoch+1}_ResNet.pth')

            if r2 > best_r2:
                best_r2 = r2
                final_val_outputs = all_outputs.to(device)
                final_val_labels = all_labels.to(device)

            epoch_time = time.time() - start_time
            sheet_para.append(["", "", "", epoch_time,
                               "Converging" if val_loss < best_val_loss else "Not Converging"])

            print(f"Stage {stage+1}, Epoch [{epoch + 1}/{num_epochs}], Train Loss: {train_loss:.4f}, "
                  f"Val Loss: {val_loss:.4f}, R2: {r2:.4f}, MAE: {mae:.4f}, RMSE: {rmse:.4f}")

        # 计算 SEMs：最后100 epoch均值 [TrainLoss, ValRMSE, ValMAE]
        sems_vec = torch.tensor([
            np.mean(train_losses),
            np.mean(val_rmses),
            np.mean(val_maes)
        ], dtype=torch.float32).to(device)
        stage_sems.append(sems_vec)

        stage_train_outputs.append(final_val_outputs)
        stage_train_labels.append(final_val_labels)

    for true_val, pred_val in zip(final_val_labels, final_val_outputs):
        sheet2.append([true_val.item(), pred_val.item()])

    sheet2.append(["Best R2", best_r2])

    workbook.save("ChainHTF_metrics.xlsx")
    workbook.close()
    workbook_para.save("ChainHTF_para.xlsx")
    workbook_para.close()


if __name__ == '__main__':
    train_model()
