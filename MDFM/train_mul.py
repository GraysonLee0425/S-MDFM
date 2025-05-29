import torch
import openpyxl
from torch import nn, optim
from torch.utils.data import DataLoader
from openpyxl import Workbook
import numpy as np
from sklearn import metrics
from Dataloder_mul import MyDataset
from model_mul import resnet18
import time

device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

img_paths = [
    "F:/wq-data/final/1-20/rgb/",
    "F:/wq-data/final/1-20/blue/",
    "F:/wq-data/final/1-20/green/",
    "F:/wq-data/final/1-20/red/",
    "F:/wq-data/final/1-20/rededge/",
    "F:/wq-data/final/1-20/nir/",
    "F:/wq-data/final/1-20/3T/"
]
label_path_train = 'Dataset/1_20/1_20_ndyield_train.json'
label_path_val = 'Dataset/1_20/1_20_ndyield_val.json'
excel_path = 'Dataset/1_20_nd-yield.csv'

TrainDataset = MyDataset(*img_paths, label_path_train, excel_path, 0)
ValDataset = MyDataset(*img_paths, label_path_val, excel_path, 1)

train_loader = DataLoader(
    dataset=TrainDataset,
    batch_size=16,
    shuffle=True,
    num_workers=4,
    pin_memory=True
)
val_loader = DataLoader(
    dataset=ValDataset,
    batch_size=8,
    shuffle=False,
    num_workers=4,
    pin_memory=True
)

model = resnet18(num_classes=1, include_top=True, text_feature_dim=18).to(device)
criterion = nn.MSELoss()
optimizer = optim.Adam(model.parameters(), lr=0.0001)
num_params = sum(p.numel() for p in model.parameters() if p.requires_grad)

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "Metrics"
sheet1.append(["Epoch", "Train Loss", "Train R2", "Train MAE", "Train RMSE", "Val Loss", "Val R2", "Val MAE", "Val RMSE"])

sheet2 = workbook.create_sheet(title="Predictions")
sheet2.append(["True Values", "Predicted Values"])

workbook_para = openpyxl.Workbook()
sheet_para = workbook_para.active
sheet_para.title = "Training Parameters"
sheet_para.append(["Batch Size (Train)", "Batch Size (Val)", "Number of Parameters", "Training Time (s)", "Convergence Status"])


best_val_loss = float('inf')
best_r2 = float('-inf')
num_epochs = 500

sheet_para.append([16, 8, num_params, "", ""])

def train_model():
    global best_val_loss, best_r2
    final_val_outputs = []
    final_val_labels = []

    for epoch in range(num_epochs):
        start_time = time.time()

        model.train()
        running_loss = 0.0
        all_train_outputs = []
        all_train_labels = []
        for inputs, text_features, labels in train_loader:

            inputs, text_features, labels = inputs.float().to(device), text_features.float().to(device), labels.float().unsqueeze(1).to(device)

            optimizer.zero_grad()
            outputs = model(inputs, text_features)
            loss = criterion(outputs, labels)
            loss.backward()
            optimizer.step()

            running_loss += loss.item() * inputs.size(0)
            all_train_outputs.append(outputs.cpu())
            all_train_labels.append(labels.cpu())

        train_loss = running_loss / len(train_loader.dataset)

        all_train_outputs = torch.cat(all_train_outputs)
        all_train_labels = torch.cat(all_train_labels)
        train_r2 = 1 - ((all_train_outputs - all_train_labels).pow(2).sum() / (all_train_labels - all_train_labels.mean()).pow(2).sum()).item()
        train_mae = torch.abs(all_train_outputs - all_train_labels).mean().item()
        train_rmse = torch.sqrt(torch.mean((all_train_outputs - all_train_labels) ** 2)).item()

        model.eval()
        val_loss = 0.0
        all_outputs = []
        all_labels = []
        with torch.no_grad():
            for inputs, text_features, labels in val_loader:

                inputs, text_features, labels = inputs.float().to(device), text_features.float().to(device), labels.float().unsqueeze(1).to(device)
                outputs = model(inputs, text_features)
                loss = criterion(outputs, labels)
                val_loss += loss.item() * inputs.size(0)
                all_outputs.append(outputs.cpu())
                all_labels.append(labels.cpu())

        val_loss = val_loss / len(val_loader.dataset)

        all_outputs = torch.cat(all_outputs)
        all_labels = torch.cat(all_labels)
        r2 = 1 - ((all_outputs - all_labels).pow(2).sum() / (all_labels - all_labels.mean()).pow(2).sum()).item()
        mae = torch.abs(all_outputs - all_labels).mean().item()
        rmse = torch.sqrt(torch.mean((all_outputs - all_labels) ** 2)).item()

        sheet1.append([epoch + 1, train_loss, train_r2, train_mae, train_rmse, val_loss, r2, mae, rmse])

        if val_loss < best_val_loss:
            best_val_loss = val_loss
            torch.save(model.state_dict(), f'1_20_ValR18HTF/{epoch + 1}_ResNet.pth')

        if r2 > best_r2:
            best_r2 = r2
            final_val_outputs = all_outputs
            final_val_labels = all_labels

        epoch_time = time.time() - start_time
        sheet_para.append(["", "", "", epoch_time, "Converging" if val_loss < best_val_loss else "Not Converging"])

        print(f"Epoch [{epoch + 1}/{num_epochs}], Train Loss: {train_loss:.4f}, Val Loss: {val_loss:.4f}, R2: {r2:.4f}, MAE: {mae:.4f}, RMSE: {rmse:.4f}")

    for true_val, pred_val in zip(final_val_labels, final_val_outputs):
        sheet2.append([true_val.item(), pred_val.item()])

    sheet2.append(["Best R2", best_r2])

    workbook.save("1_20_ValR18HTF_metrics.xlsx")
    workbook.close()
    workbook_para.save("1_20_ValR18HTF_para.xlsx")
    workbook_para.close()

if __name__ == '__main__':
    train_model()
